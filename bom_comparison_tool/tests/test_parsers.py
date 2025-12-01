import pytest
import os
from openpyxl import Workbook
from bom_comparison_tool.core.parsers import parse_bom_file
from bom_comparison_tool.core.models import BOMItem, ErrorDict

# Fixture to create a dummy XLSX file for testing
@pytest.fixture
def dummy_xlsx_file(tmp_path):
    file_path = tmp_path / "test_master.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    
    # Write header
    ws['A1'] = "MPN"
    ws['B1'] = "Quantity"
    ws['C1'] = "RefDes"
    ws['D1'] = "Description"
    
    # Write data
    ws['A2'] = "PART-001"
    ws['B2'] = 10
    ws['C2'] = "R1, R2"
    ws['D2'] = "Resistor 10k"
    
    ws['A3'] = "PART-002"
    ws['B3'] = 5
    ws['C3'] = "C1"
    ws['D3'] = "Capacitor 100nF"

    wb.save(file_path)
    return file_path

# Test case for XLSX parser
def test_parse_xlsx_success(dummy_xlsx_file):
    result = parse_bom_file(str(dummy_xlsx_file))
    
    assert 'error' not in result # Check if it's not an ErrorDict
    assert isinstance(result, list)
    assert len(result) == 2
    
    # Check first item
    assert result[0]['MPN'] == "PART-001"
    assert result[0]['Quantity'] == 10
    assert result[0]['RefDes'] == ["R1", "R2"]
    assert result[0]['Description'] == "Resistor 10k"
    
    # Check second item
    assert result[1]['MPN'] == "PART-002"
    assert result[1]['Quantity'] == 5
    assert result[1]['RefDes'] == ["C1"]
    assert result[1]['Description'] == "Capacitor 100nF"

# Test case for unsupported file type
def test_parse_unsupported_file(tmp_path):
    file_path = tmp_path / "unsupported.xyz"
    file_path.write_text("dummy content")
    
    result = parse_bom_file(str(file_path))
    
    assert 'error' in result # Check if it is an ErrorDict
    assert "Unsupported file extension" in result['error']

# Test case for file not found
def test_parse_file_not_found():
    result = parse_bom_file("non_existent_file.xlsx")
    
    assert 'error' in result # Check if it is an ErrorDict
    assert "File not found" in result['error']
